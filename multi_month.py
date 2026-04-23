"""
MULTI-MONTH ANALYSIS MODULE
============================
Adds the Monthly Analysis tab to the expense tracker.

Features:
- Month-by-month bar chart
- Category × Month table where every cell is a clickable button
- Clicking a cell expands a panel showing every transaction inside it
- Each transaction has a dropdown to reassign its category
- Changes apply immediately to all charts and totals
- Attention flags for unusual spending spikes
"""

import streamlit as st
import pandas as pd
import plotly.express as px


# ── 10 fixed categories ───────────────────────────────────────────────────────
ALL_CATEGORIES = [
    'Rent',
    'Family',
    'Food',
    'Travel',
    'Medical',
    'Subscriptions & Devices',
    'Books',
    'Garden',
    'Gifts',
    'Miscellaneous',
]

CATEGORY_COLORS = {
    'Rent':                    '#795548',
    'Family':                  '#00BCD4',
    'Food':                    '#4CAF50',
    'Travel':                  '#2196F3',
    'Medical':                 '#F44336',
    'Subscriptions & Devices': '#3F51B5',
    'Books':                   '#9C27B0',
    'Garden':                  '#8BC34A',
    'Gifts':                   '#E91E63',
    'Miscellaneous':           '#9E9E9E',
}


# ============================================================================
# DATA HELPERS
# ============================================================================

def get_sorted_months(df):
    """Return month labels in chronological order e.g. ['Jan 2026', 'Feb 2026']"""
    df = df.copy()
    df['_period'] = df['Date'].dt.to_period('M')
    df['_label']  = df['Date'].dt.strftime('%b %Y')
    mapping = df[['_period', '_label']].drop_duplicates().sort_values('_period')
    return mapping['_label'].tolist()


def build_monthly_totals(df, months):
    """Total spending per month as a Series."""
    df = df.copy()
    df['Month'] = df['Date'].dt.strftime('%b %Y')
    totals = df.groupby('Month')['Amount'].sum()
    return totals.reindex(months, fill_value=0)


def build_pivot_table(df, months):
    """
    Category x Month pivot table.
    Rows = categories, Columns = months, Values = total spent.
    Sorted by row total descending.
    """
    df = df.copy()
    df['Month'] = df['Date'].dt.strftime('%b %Y')
    pivot = df.pivot_table(
        index='Category',
        columns='Month',
        values='Amount',
        aggfunc='sum',
        fill_value=0
    )
    pivot = pivot.reindex(columns=months, fill_value=0)
    pivot['Total'] = pivot[months].sum(axis=1)
    return pivot.sort_values('Total', ascending=False)


def get_attention_flags(pivot, months, threshold=0.30):
    """
    Return list of dicts for categories where any month is
    30%+ above that category's own average across months.
    """
    flags = []
    for category in pivot.index:
        vals = pivot.loc[category, months]
        avg  = vals.mean()
        if avg == 0:
            continue
        for month in months:
            val = pivot.loc[category, month]
            if val > avg * (1 + threshold):
                flags.append({
                    'Category': category,
                    'Month':    month,
                    'Amount':   val,
                    'Avg':      round(avg, 0),
                    'PctAbove': int(((val - avg) / avg) * 100),
                })
    return sorted(flags, key=lambda x: x['PctAbove'], reverse=True)


def _apply_edits(df):
    """Apply any category edits stored in session state to the dataframe."""
    df = df.copy()
    edits = st.session_state.get('edited_categories', {})
    for idx, new_cat in edits.items():
        if idx in df.index:
            df.loc[idx, 'Category'] = new_cat
    return df



# ============================================================================
# REPORT GENERATORS
# ============================================================================

def generate_monthly_excel(df, months, pivot):
    """
    Generate an Excel file with:
    - Sheet 1: Category x Month summary table
    - One sheet per month with all transactions for that month
    Returns bytes ready for st.download_button.
    """
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    df = df.copy()
    df['Month'] = df['Date'].dt.strftime('%b %Y')

    output = BytesIO()
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"

    header_fill = PatternFill("solid", fgColor="2196F3")
    header_font = Font(color="FFFFFF", bold=True)
    bold = Font(bold=True)

    # Write header
    ws.cell(1, 1, "Category").font = header_font
    ws.cell(1, 1).fill = header_fill
    for i, month in enumerate(months):
        cell = ws.cell(1, i + 2, month)
        cell.font = header_font
        cell.fill = header_fill
    total_col = len(months) + 2
    ws.cell(1, total_col, "Total").font = header_font
    ws.cell(1, total_col).fill = header_fill

    # Write category rows
    for r, category in enumerate(pivot.index, start=2):
        ws.cell(r, 1, category)
        for i, month in enumerate(months):
            val = pivot.loc[category, month]
            ws.cell(r, i + 2, round(val, 0))
        ws.cell(r, total_col, round(pivot.loc[category, 'Total'], 0))

    # Totals row
    total_row = len(pivot) + 2
    ws.cell(total_row, 1, "TOTAL").font = bold
    for i, month in enumerate(months):
        ws.cell(total_row, i + 2, round(pivot[month].sum(), 0)).font = bold
    ws.cell(total_row, total_col, round(pivot['Total'].sum(), 0)).font = bold

    # Column widths
    ws.column_dimensions['A'].width = 26
    for i in range(len(months) + 1):
        ws.column_dimensions[get_column_letter(i + 2)].width = 14

    # ── One sheet per month ────────────────────────────────────────────────────
    for month in months:
        month_df = df[df['Month'] == month].sort_values('Date')
        ws2 = wb.create_sheet(title=month)

        # Header
        headers = ['Date', 'Description', 'Type', 'Category', 'Amount']
        for i, h in enumerate(headers, 1):
            cell = ws2.cell(1, i, h)
            cell.font = header_font
            cell.fill = header_fill

        # Transactions
        for r, (_, row) in enumerate(month_df.iterrows(), start=2):
            parts = str(row['Particulars']).split('/')
            note = parts[3].strip() if len(parts) > 3 and parts[3].strip() not in ['','0000'] else ''
            desc = note if note else (parts[2].strip() if len(parts) > 2 else str(row['Particulars'])[:40])
            ws2.cell(r, 1, row['Date'].strftime('%d %b %Y'))
            ws2.cell(r, 2, desc[:45])
            ws2.cell(r, 3, str(row['Transaction Type']).strip())
            ws2.cell(r, 4, row['Category'])
            ws2.cell(r, 5, round(row['Amount'], 0))

        # Total row
        total_r = len(month_df) + 2
        ws2.cell(total_r, 4, "TOTAL").font = bold
        ws2.cell(total_r, 5, round(month_df['Amount'].sum(), 0)).font = bold

        # Widths
        ws2.column_dimensions['A'].width = 14
        ws2.column_dimensions['B'].width = 32
        ws2.column_dimensions['C'].width = 10
        ws2.column_dimensions['D'].width = 24
        ws2.column_dimensions['E'].width = 12

    wb.save(output)
    return output.getvalue()


def generate_monthly_pdf(df, months, pivot, flags):
    """
    Generate a PDF summary report with:
    - Category x Month table
    - Monthly totals
    - Attention flags
    Returns bytes ready for st.download_button.
    """
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph,
        Spacer, HRFlowable
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from datetime import datetime

    output = BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=A4,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm
    )

    styles = getSampleStyleSheet()
    title_style  = ParagraphStyle('title',  fontSize=16, fontName='Helvetica-Bold', spaceAfter=6)
    sub_style    = ParagraphStyle('sub',    fontSize=11, fontName='Helvetica-Bold', spaceAfter=4, spaceBefore=12)
    caption_style= ParagraphStyle('cap',    fontSize=8,  fontName='Helvetica',      textColor=colors.grey, spaceAfter=6)
    flag_style   = ParagraphStyle('flag',   fontSize=9,  fontName='Helvetica',      spaceAfter=4)

    blue  = colors.HexColor('#2196F3')
    light = colors.HexColor('#E3F2FD')
    red   = colors.HexColor('#F44336')
    orange= colors.HexColor('#FF9800')
    amber = colors.HexColor('#FFC107')

    story = []

    # Title
    story.append(Paragraph("Monthly Expense Report", title_style))
    period = f"{months[0]} — {months[-1]}"
    story.append(Paragraph(f"{period}  ·  Generated {datetime.now().strftime('%d %b %Y')}", caption_style))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.lightgrey, spaceAfter=10))

    # ── Summary table ──────────────────────────────────────────────────────────
    story.append(Paragraph("Category Breakdown", sub_style))

    header = ['Category'] + months + ['Total']
    table_data = [header]
    for category in pivot.index:
        row = [category]
        for month in months:
            val = pivot.loc[category, month]
            row.append(f"Rs{val:,.0f}" if val > 0 else "—")
        row.append(f"Rs{pivot.loc[category, 'Total']:,.0f}")
        table_data.append(row)

    # Grand total row
    grand = ['TOTAL']
    for month in months:
        grand.append(f"Rs{pivot[month].sum():,.0f}")
    grand.append(f"Rs{pivot['Total'].sum():,.0f}")
    table_data.append(grand)

    col_widths = [4.5*cm] + [2.8*cm] * len(months) + [3*cm]
    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND',   (0, 0), (-1, 0),              blue),
        ('TEXTCOLOR',    (0, 0), (-1, 0),              colors.white),
        ('FONTNAME',     (0, 0), (-1, 0),              'Helvetica-Bold'),
        ('FONTSIZE',     (0, 0), (-1, -1),             8),
        ('ALIGN',        (1, 0), (-1, -1),             'RIGHT'),
        ('ALIGN',        (0, 0), (0, -1),              'LEFT'),
        ('BACKGROUND',   (0, -1), (-1, -1),            light),
        ('FONTNAME',     (0, -1), (-1, -1),            'Helvetica-Bold'),
        ('ROWBACKGROUNDS',(0, 1), (-1, -2),            [colors.white, colors.HexColor('#F5F5F5')]),
        ('GRID',         (0, 0), (-1, -1),             0.3, colors.lightgrey),
        ('TOPPADDING',   (0, 0), (-1, -1),             4),
        ('BOTTOMPADDING',(0, 0), (-1, -1),             4),
        ('LEFTPADDING',  (0, 0), (-1, -1),             6),
        ('RIGHTPADDING', (0, 0), (-1, -1),             6),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 0.4*cm))

    # ── Attention flags ───────────────────────────────────────────────────────
    if flags:
        story.append(Paragraph("Attention Needed", sub_style))
        story.append(Paragraph(
            "Categories where any month is 30%+ above that category's average.",
            caption_style
        ))

        flag_data = [['Category', 'Month', 'Spent', 'Average', 'Above avg']]
        for flag in flags:
            pct = flag['PctAbove']
            flag_data.append([
                flag['Category'],
                flag['Month'],
                f"Rs{flag['Amount']:,.0f}",
                f"Rs{flag['Avg']:,.0f}",
                f"+{pct}%",
            ])

        ftbl = Table(flag_data, colWidths=[4.5*cm, 2.5*cm, 2.8*cm, 2.8*cm, 2.5*cm], repeatRows=1)
        ftbl.setStyle(TableStyle([
            ('BACKGROUND',   (0, 0), (-1, 0),   blue),
            ('TEXTCOLOR',    (0, 0), (-1, 0),   colors.white),
            ('FONTNAME',     (0, 0), (-1, 0),   'Helvetica-Bold'),
            ('FONTSIZE',     (0, 0), (-1, -1),  8),
            ('ALIGN',        (2, 0), (-1, -1),  'RIGHT'),
            ('GRID',         (0, 0), (-1, -1),  0.3, colors.lightgrey),
            ('ROWBACKGROUNDS',(0,1), (-1,-1),   [colors.white, colors.HexColor('#FFF8E1')]),
            ('TOPPADDING',   (0, 0), (-1, -1),  4),
            ('BOTTOMPADDING',(0, 0), (-1, -1),  4),
            ('LEFTPADDING',  (0, 0), (-1, -1),  6),
            ('RIGHTPADDING', (0, 0), (-1, -1),  6),
        ]))
        story.append(ftbl)

    doc.build(story)
    return output.getvalue()

# ============================================================================
# TRANSACTION DRILL-DOWN PANEL
# ============================================================================

def render_drilldown(df, category, month):
    """
    Show every transaction inside one Category x Month cell.
    Each row has a dropdown to reassign the category.
    Changes are stored immediately in st.session_state.edited_categories.
    """
    df = df.copy()
    df['Month'] = df['Date'].dt.strftime('%b %Y')

    # Filter to this cell
    mask = (df['Category'] == category) & (df['Month'] == month)
    rows = df[mask].sort_values('Date')
    total = rows['Amount'].sum()

    # Panel header
    st.markdown(
        f"<div style='background:var(--background-secondary,#f8f8f8);"
        f"padding:12px 16px;border-radius:8px;margin-bottom:8px'>"
        f"<b style='font-size:16px'>{category} · {month}</b>"
        f"&nbsp;&nbsp;<span style='color:grey'>"
        f"{len(rows)} transactions · ₹{total:,.0f} total</span></div>",
        unsafe_allow_html=True
    )
    st.caption("Change a category using the dropdown — the table above updates immediately.")

    if rows.empty:
        st.info("No transactions in this cell.")
        return

    # Column headers
    h1, h2, h3, h4 = st.columns([1, 4, 2, 3])
    h1.markdown("**Date**")
    h2.markdown("**Description**")
    h3.markdown("**Amount**")
    h4.markdown("**Category**")

    changed = False

    for idx, row in rows.iterrows():
        c1, c2, c3, c4 = st.columns([1, 4, 2, 3])

        # Date
        c1.write(row['Date'].strftime('%d %b'))

        # Description — prefer the human-written UPI note over the full string
        parts = str(row['Particulars']).split('/')
        note = parts[3].strip() if len(parts) > 3 else ''
        if note and note.lower() not in ['0000', 'upi', 'up', 'f', 't', '']:
            desc = note
        else:
            # Fall back to merchant name (3rd segment of UPI string)
            merchant = parts[2].strip() if len(parts) > 2 else ''
            desc = merchant if merchant else str(row['Particulars'])[:35]
        c2.write(desc[:45])

        # Amount
        c3.write(f"₹{row['Amount']:,.0f}")

        # Category dropdown — pre-selected to current category
        current = st.session_state.get('edited_categories', {}).get(idx, row['Category'])
        current_idx = ALL_CATEGORIES.index(current) if current in ALL_CATEGORIES else 0

        new_cat = c4.selectbox(
            label="category",
            options=ALL_CATEGORIES,
            index=current_idx,
            key=f"dd_{idx}",
            label_visibility="collapsed",
        )

        # Detect change and save
        prev = st.session_state.get('edited_categories', {}).get(idx, row['Category'])
        if new_cat != prev:
            if 'edited_categories' not in st.session_state:
                st.session_state.edited_categories = {}
            st.session_state.edited_categories[idx] = new_cat
            changed = True

    if changed:
        st.rerun()

    # Close button
    st.markdown("")
    if st.button("✕ Close breakdown", key=f"close_{category}_{month}"):
        st.session_state.drilldown_key = None
        st.rerun()


# ============================================================================
# MAIN TAB RENDERER
# ============================================================================

def render_monthly_analysis_tab(df):
    """
    Entry point called from app.py to render the Monthly Analysis tab.
    Expects df with columns: Date, Particulars, Transaction Type, Category, Amount.
    """
    st.header("Monthly Analysis")

    if df is None or len(df) == 0:
        st.info("No data loaded.")
        return

    months = get_sorted_months(df)
    if len(months) < 2:
        st.info("Upload a file with at least 2 months of data to see comparison.")
        return

    # Apply any category edits from this session
    df = _apply_edits(df)

    # Initialise drilldown state
    if 'drilldown_key' not in st.session_state:
        st.session_state.drilldown_key = None

    # ── Section 1: Monthly totals bar chart ──────────────────────────────────
    st.subheader("Total spending by month")
    monthly_totals = build_monthly_totals(df, months)

    fig = px.bar(
        x=monthly_totals.index,
        y=monthly_totals.values,
        labels={'x': '', 'y': 'Amount (₹)'},
        color=monthly_totals.index,
        color_discrete_sequence=[
            '#2196F3', '#4CAF50', '#FF9800', '#E91E63',
            '#9C27B0', '#00BCD4', '#F44336', '#795548',
        ],
        text=monthly_totals.values,
    )
    fig.update_traces(texttemplate='₹%{text:,.0f}', textposition='outside')
    fig.update_layout(
        showlegend=False, height=320,
        margin=dict(t=40, b=10, l=10, r=10),
    )
    st.plotly_chart(fig, use_container_width=True)

    st.markdown("---")

    # ── Section 2: Category × Month table ────────────────────────────────────
    st.subheader("Click any amount to see and edit its transactions")

    pivot = build_pivot_table(df, months)

    # Table header row
    n = len(months)
    header_cols = st.columns([3] + [2] * n + [2])
    header_cols[0].markdown("**Category**")
    for i, m in enumerate(months):
        header_cols[i + 1].markdown(f"**{m}**")
    header_cols[-1].markdown("**Total**")

    st.markdown(
        "<hr style='margin:4px 0 8px 0;border:none;border-top:1px solid #ddd'>",
        unsafe_allow_html=True
    )

    # One row per category
    for category in pivot.index:
        row_cols = st.columns([3] + [2] * n + [2])

        # Category label with colour dot
        color = CATEGORY_COLORS.get(category, '#9E9E9E')
        row_cols[0].markdown(
            f"<span style='color:{color};font-size:16px'>●</span>&nbsp;{category}",
            unsafe_allow_html=True
        )

        # Month cells — each is a button if value > 0
        for i, month in enumerate(months):
            val = pivot.loc[category, month]
            cell_key = f"{category}||{month}"
            is_open  = st.session_state.drilldown_key == cell_key

            if val > 0:
                if row_cols[i + 1].button(
                    f"₹{val:,.0f}",
                    key=f"btn_{category}_{month}",
                    use_container_width=True,
                    type="primary" if is_open else "secondary",
                ):
                    st.session_state.drilldown_key = None if is_open else cell_key
                    st.rerun()
            else:
                row_cols[i + 1].markdown(
                    "<span style='color:#bbb'>—</span>", unsafe_allow_html=True
                )

        # Row total (not clickable)
        row_cols[-1].markdown(f"**₹{pivot.loc[category, 'Total']:,.0f}**")

        # Drilldown panel opens directly below the row it belongs to
        for month in months:
            if st.session_state.drilldown_key == f"{category}||{month}":
                with st.container():
                    render_drilldown(df, category, month)
                    st.markdown(
                        "<hr style='margin:8px 0;border:none;border-top:1px solid #eee'>",
                        unsafe_allow_html=True
                    )

    st.markdown("---")

    # ── Section 3: Attention flags ────────────────────────────────────────────
    st.subheader("Attention needed")
    st.caption("Categories where any month is 30%+ above that category's own average")

    flags = get_attention_flags(pivot, months)

    if not flags:
        st.success("✅ No unusual spikes — spending is consistent across months.")
    else:
        cols = st.columns(2)
    for i, flag in enumerate(flags):
        pct = flag['PctAbove']
        if pct >= 80:
            color, emoji = '#F44336', '🔴'
        elif pct >= 50:
            color, emoji = '#FF9800', '🟠'
        else:
            color, emoji = '#FFC107', '🟡'

        with cols[i % 2]:
            st.markdown(
                f"<div style='border-left:4px solid {color};"
                f"padding:10px 14px;margin-bottom:10px;"
                f"border-radius:0 8px 8px 0'>"
                f"<b>{emoji} {flag['Category']} — {flag['Month']}</b><br>"
                f"₹{flag['Amount']:,.0f} vs avg ₹{flag['Avg']:,.0f} "
                f"<span style='color:{color};font-weight:600'>+{pct}%</span>"
                f"</div>",
                unsafe_allow_html=True
            )
            cell_key = f"{flag['Category']}||{flag['Month']}"
            if st.button(
                "See transactions →",
                key=f"flag_{flag['Category']}_{flag['Month']}",
            ):
                st.session_state.drilldown_key = cell_key
                st.rerun()

    st.markdown("---")

    # ── Section 4: Download monthly report ────────────────────────────────────
    st.subheader("Download monthly report")

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("**Excel — full breakdown**")
        st.caption("One sheet per month with all transactions")
        excel_data = generate_monthly_excel(df, months, pivot)
        st.download_button(
            label="Download Excel",
            data=excel_data,
            file_name=f"monthly_report_{months[0].replace(' ','_')}_{months[-1].replace(' ','_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    with col2:
        st.markdown("**PDF — summary report**")
        st.caption("Category totals, bar chart, attention flags")
        if st.button("Generate PDF", use_container_width=True, type="secondary"):
            with st.spinner("Creating PDF..."):
                pdf_data = generate_monthly_pdf(df, months, pivot, flags)
                st.session_state.monthly_pdf = pdf_data
                st.session_state.monthly_pdf_name = f"monthly_report_{months[0].replace(' ','_')}_{months[-1].replace(' ','_')}.pdf"

        if 'monthly_pdf' in st.session_state:
            st.download_button(
                label="Download PDF",
                data=st.session_state.monthly_pdf,
                file_name=st.session_state.monthly_pdf_name,
                mime="application/pdf",
                use_container_width=True,
            )

